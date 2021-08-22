import datetime

import openpyxl

# def getRowCount (file,sheetName):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook.get_sheet_by_name(sheetName)
#     return (sheet.max_row)
#
# def getColumnCount(file,sheetName):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook.get_sheet_by_name(sheetName)
#     return (sheet.max_column)
#
# def readData(file,sheetName,rownum,columnno):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook.get_sheet_by_name(sheetName)
#     return sheet.cell(row=rownum, column=columnno).value
#
# def writeData(file,sheetName,rownum,columnno,data):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook.get_sheet_by_name(sheetName)
#     sheet.cell(row=rownum, column=columnno).value=data
#     workbook.save(file)
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pandas import read_excel


def getData(path, sheetName):
    data = read_excel(path,sheet_name=sheetName)
    return data
def updateResult(path,sheetName,columnName,passResult,rowNum,color):
    df = pd.DataFrame({columnName:[passResult]})
    wb = load_workbook(path)
    ws = wb[sheetName]

    redFill = PatternFill(start_color='00FF0000',fill_type='solid')
    greenFill = PatternFill(start_color='00FF0000',fill_type='solid')

    for index, row in df.iterrows():
        cell='c%d' % (index+2+rowNum)
        ws[cell] = row[columnName]
        if color=='Red':
            ws[cell].fill = redFill
        elif color== 'Green':
            ws[cell].fill = greenFill
    wb.save(path)

def updateExecutionCompletion(path,sheetName,rowNum):

    df = pd.DataFrame({'ExecutionTime':[str(datetime.datetime.now())]})
    wb = load_workbook(path)
    ws = wb[sheetName]

    for index, row in df.iterrows():
        cell = 'D%d' % (index+2+rowNum)
        ws[cell] = row['ExecutionTime']
    wb.save()